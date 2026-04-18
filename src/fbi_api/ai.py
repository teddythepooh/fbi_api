import json
import os
import re
from datetime import datetime

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_RESULT_COL_ORDER = [
    "ori", "offense", "year", "month", "observed_count",
    "expected_range", "flag", "explanation", "sources",
]

_FLAG_COLORS = {
    "EXPLAINABLE":    "C6EFCE",
    "POSSIBLE_ERROR": "FFEB9C",
    "UNEXPLAINED":    "FFC7CE",
}

_NUM_COLS   = 8

_COL_WIDTHS = {1: 20, 2: 8, 3: 8, 4: 16, 5: 28, 6: 16, 7: 60, 8: 50}

_HEADERS    = [
    "Offense", "Year", "Month", "Observed Count",
    "Expected Range", "Flag", "Explanation", "Sources",
]

_HEADER_FILL = PatternFill("solid", fgColor = "2F4F8F")

_THIN_BORDER = Border(
    left   = Side(style = "thin", color = "CCCCCC"),
    right  = Side(style = "thin", color = "CCCCCC"),
    top    = Side(style = "thin", color = "CCCCCC"),
    bottom = Side(style = "thin", color = "CCCCCC"),
)

_WRAP = Alignment(wrap_text = True, vertical = "top")

class AnomalyDetection:
    def __init__(self, anthropic_api_key: str = None):
        '''
        anthropic_api_key: If not passed, ANTHROPIC_API_KEY is automatically invoked —
        consistent with how the FBI class handles FBI_API_KEY.

        Instantiates the AnomalyDetection client. Requires the 'anthropic' package:
        pip install fbi-data-api[audit]
        '''
        self.client = anthropic.Anthropic(
            api_key = anthropic_api_key or os.environ.get("ANTHROPIC_API_KEY")
        )

    def flag_anomalies_with_ai(
        self,
        df: pd.DataFrame,
        max_searches: int = 3,
        verbose: bool = True,
        output_path: str = None,
    ) -> pd.DataFrame:
        wb            = Workbook()
        wb.remove(wb.active)
        all_rows      = []
        ori_anomalies = {ori: [] for ori in df["ori"].unique()}

        metadata = self._fetch_metadata(df)

        for (ori, offense), group_df in df.groupby(["ori", "offense"]):
            years       = sorted(group_df["year"].unique().tolist())
            agency_info = self._get_agency_info(metadata, ori)
            parsed      = self._call_claude(group_df, ori, offense, years, max_searches, agency_info)
            anomalies = parsed.get("anomalies", [])
            summary   = parsed.get("summary", "")

            if verbose:
                self._print_group_summary(ori, offense, years, anomalies, summary)

            for a in anomalies:
                a["ori"]     = ori
                a["offense"] = offense
                all_rows.append(a)

            ori_anomalies[ori].extend(anomalies)

        for ori, anomalies in ori_anomalies.items():
            self._add_ori_sheet(wb, ori, anomalies)

        result_df = self._build_result_df(all_rows)

        path = output_path or self._default_output_path(df)
        wb.save(path)
        print(f"\nReport exported to: {path}")

        return result_df

    def _fetch_metadata(self, df: pd.DataFrame) -> pd.DataFrame:
        from fbi_api import FBI
        fbi         = FBI()
        state_abbrs = df["ori"].str[:2].unique().tolist()
        frames      = [fbi.get_metadata(state_abbr = s) for s in state_abbrs]
        return pd.concat(frames, ignore_index = True)

    @staticmethod
    def _get_agency_info(metadata: pd.DataFrame, ori: str) -> dict:
        row = metadata[metadata["ori"] == ori]
        if row.empty:
            return {}
        row = row.iloc[0]
        return {
            "agency_name"      : row.get("agency_name", ""),
            "agency_type_name" : row.get("agency_type_name", ""),
            "counties"         : row.get("counties", ""),
            "state_name"       : row.get("state_name", ""),
            "state_abbr"       : row.get("state_abbr", ""),
            "latitude"         : row.get("latitude", ""),
            "longitude"        : row.get("longitude", ""),
            "is_nibrs"         : row.get("is_nibrs", ""),
            "nibrs_start_date" : row.get("nibrs_start_date", "N/A"),
        }

    @staticmethod
    def _build_result_df(all_rows: list) -> pd.DataFrame:
        if not all_rows:
            return pd.DataFrame(columns = _RESULT_COL_ORDER)

        result_df            = pd.DataFrame(all_rows)
        result_df["sources"] = result_df["sources"].apply(
            lambda x: "; ".join(x) if isinstance(x, list) else x
        )
        return result_df[[c for c in _RESULT_COL_ORDER if c in result_df.columns]]

    @staticmethod
    def _default_output_path(df: pd.DataFrame) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        oris      = "_".join(sorted(df["ori"].unique().tolist()))
        if len(oris) > 40:
            oris = oris[:37] + "..."
        return f"fbi_anomalies_{oris}_{timestamp}.xlsx"

    def _call_claude(
        self,
        group_df: pd.DataFrame,
        ori: str,
        offense: str,
        years: list,
        max_searches: int,
        agency_info: dict = {},
    ) -> dict:
        prompt   = self._build_prompt(group_df, ori, offense, years, agency_info)
        response = self.client.messages.create(
            model      = "claude-sonnet-4-6",
            max_tokens = 2000,
            tools      = [
                {
                    "type"     : "web_search_20250305",
                    "name"     : "web_search",
                    "max_uses" : max_searches,
                },
            ],
            messages = [{"role": "user", "content": prompt}],
        )
        return self._parse_claude_response(response, ori, offense)

    @staticmethod
    def _build_prompt(
        group_df: pd.DataFrame,
        ori: str,
        offense: str,
        years: list,
        agency_info: dict = {},
    ) -> str:
        data_str = (
            group_df[["year", "month", "count"]]
            .sort_values(["year", "month"])
            .to_string(index = False)
        )
        if agency_info:
            agency_block = (
                f"Agency name:  {agency_info.get('agency_name', 'unknown')}\n"
                f"Agency type:  {agency_info.get('agency_type_name', 'unknown')}\n"
                f"County:       {agency_info.get('counties', 'unknown')}\n"
                f"State:        {agency_info.get('state_name', 'unknown')} "
                f"({agency_info.get('state_abbr', 'unknown')})\n"
                f"Coordinates:  {agency_info.get('latitude', 'unknown')}, "
                f"{agency_info.get('longitude', 'unknown')}\n"
                f"NIBRS:        {'Yes' if agency_info.get('is_nibrs') else 'No'} "
                f"(started: {agency_info.get('nibrs_start_date', 'N/A')})"
            )
        else:
            agency_block = "Agency metadata unavailable."

        return f"""
Assume the role of a criminilogist who is auditing monthly crime statistics submitted to the FBI's Uniform Crime Reporting program.

ORI:     {ori}
Offense: {offense}
Years:   {", ".join(str(y) for y in years)}

Agency context:
{agency_block}
 
IMPORTANT: Do not use the data below to form your expectations. The data is what you are auditingm, so it may be wrong. Your reference point must come from external sources found via web search, not from patterns within this dataset.
 
Monthly counts to audit:
 
{data_str}
 
Your task has two steps. Complete both before responding.
 
Step 1 — Search first. Search the web for:
- What the true or expected level of {offense} should be for this agency during these years, based on external reporting, news coverage, or published crime statistics.
- Known reporting issues, undercounting, non-participation, or data quality problems for this specific agency (ORI: {ori}) and offense type.
- Any transitions between reporting systems (e.g. SRS to NIBRS), reclassification of offenses, or changes in how this agency counts this crime.
- Audits, inspector general reports, academic studies, or journalism raising concerns about this agency's crime statistics.
- Real-world events (surges, policy changes, major incidents) that would be expected to produce a visible shift in counts.
 
Step 2 — Audit the data. Using only external sources as your benchmark:
- Flag every month where the reported count appears inconsistent with what external sources suggest the true count should be. Do not limit yourself to statistical outliers within the series — if the entire year looks suspect based on external knowledge, flag every month.
- Flag months where a known real-world event should have produced a visible effect but does not appear in the data.
- Flag months with implausibly low counts (including zeros) if external sources indicate this offense was actively occurring during that period.
 
Classification rules:
- EXPLAINABLE:    a confirmed real-world cause or reporting change was found via web search that fully accounts for the anomaly.
- POSSIBLE_ERROR: the counts appear inconsistent with external sources or context, suggesting a reporting problem, even if you cannot confirm a specific cause. Use this — not UNEXPLAINED — whenever the data looks implausibly low or inconsistent with known reality and no confirmed explanation exists.
- UNEXPLAINED:    the counts are statistically unusual relative to the time series AND no external context was found. Only use this if you found no relevant external information whatsoever.
 
For expected_range: derive this from external sources (news reports, published statistics, comparable jurisdictions), not from the data provided. If no external benchmark exists, write "unknown".
 
Respond only with a JSON object in this exact format and absolutely nothing else:
{{
  "anomalies": [
    {{
      "year": 2021,
      "month": 8,
      "observed_count": 142,
      "expected_range": "180-220 based on [external source]",
      "flag": "POSSIBLE_ERROR",
      "explanation": "Detailed explanation referencing what external sources indicate the true count should be and why the reported count is suspect.",
      "sources": ["https://source1.com", "https://source2.com"]
    }}
  ],
  "summary": "Overall assessment of data quality, including known reporting issues found and whether the full dataset or specific months appear suspect."
}}
 
If external sources confirm the data is accurate and no issues are found, return an empty anomalies list with a summary explaining why the data appears clean.
"""

    @staticmethod
    def _parse_claude_response(response, ori: str, offense: str) -> dict:
        result_text = "".join(
            block.text for block in response.content
            if block.type == "text"
        )

        result_text = re.sub(r"```(?:json)?\s*", "", result_text).strip()

        match = re.search(r"\{.*\}", result_text, re.DOTALL)
        if not match:
            raise ValueError(
                f"Could not parse a JSON object from Claude's response "
                f"for ORI={ori}, offense={offense}:\n{result_text}"
            )

        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            matches = list(re.finditer(r"\{.*\}", result_text, re.DOTALL))
            for m in reversed(matches):
                try:
                    return json.loads(m.group())
                except json.JSONDecodeError:
                    continue

            raise ValueError(
                f"Could not parse a JSON object from Claude's response "
                f"for ORI={ori}, offense={offense}:\n{result_text}"
            )

    @staticmethod
    def _print_group_summary(
        ori: str,
        offense: str,
        years: list,
        anomalies: list,
        summary: str,
    ) -> None:
        print(f"\nAnalysing  ORI: {ori}  |  Offense: {offense}")
        print(f"Years: {', '.join(str(y) for y in years)}")

        if not anomalies:
            print("  No anomalies detected.")
        else:
            for a in anomalies:
                print(f"\n  [{a['flag']}] {int(a['year'])}-{int(a['month']):02d}")
                print(f"    Observed : {a['observed_count']}")
                print(f"    Expected : {a.get('expected_range', 'N/A')}")
                print(f"    {a['explanation']}")
                for src in a.get("sources", []):
                    print(f"      - {src}")

        print(f"\n  Summary: {summary}")

    def _add_ori_sheet(self, wb: Workbook, ori: str, anomalies: list) -> None:
        ws = wb.create_sheet(title = ori)
        self._write_header_row(ws)
        self._write_data_rows(ws, anomalies)
        self._set_column_widths(ws)

    @staticmethod
    def _write_header_row(ws) -> None:
        header_row = ws.max_row + 1
        for col_idx, header in enumerate(_HEADERS, start = 1):
            cell           = ws.cell(row = header_row, column = col_idx, value = header)
            cell.font      = Font(name = "Arial", size = 10, bold = True, color = "FFFFFF")
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal = "center", vertical = "center")
            cell.border    = _THIN_BORDER

        ws.row_dimensions[header_row].height = 20
        ws.freeze_panes = ws.cell(row = header_row + 1, column = 1)

    @staticmethod
    def _write_data_rows(ws, anomalies: list) -> None:
        if not anomalies:
            no_data_row    = ws.max_row + 1
            cell           = ws.cell(row = no_data_row, column = 1, value = "No anomalies detected.")
            cell.font      = Font(name = "Arial", size = 10, italic = True)
            cell.alignment = _WRAP
            ws.merge_cells(
                start_row    = no_data_row, start_column = 1,
                end_row      = no_data_row, end_column   = _NUM_COLS,
            )
            return

        for a in anomalies:
            data_row    = ws.max_row + 1
            sources     = a.get("sources", [])
            explanation = a.get("explanation", "")
            flag        = a.get("flag", "")
            sources_str = "\n".join(f"{i + 1}. {src}" for i, src in enumerate(sources))
            flag_fill   = PatternFill("solid", fgColor = _FLAG_COLORS.get(flag, "FFFFFF"))

            values = [
                a.get("offense", ""),
                int(a["year"]),
                int(a["month"]),
                a["observed_count"],
                a.get("expected_range", ""),
                flag,
                explanation,
                sources_str,
            ]

            for col_idx, value in enumerate(values, start = 1):
                cell           = ws.cell(row = data_row, column = col_idx, value = value)
                cell.font      = Font(name = "Arial", size = 10)
                cell.fill      = flag_fill
                cell.border    = _THIN_BORDER
                cell.alignment = _WRAP

            explanation_lines                  = max(1, len(explanation) // 80 + explanation.count("\n"))
            ws.row_dimensions[data_row].height = max(explanation_lines, len(sources), 2) * 15

    @staticmethod
    def _set_column_widths(ws) -> None:
        for col_idx, width in _COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width